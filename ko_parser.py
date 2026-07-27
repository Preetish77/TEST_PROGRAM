import csv
import io
import re

try:
    import openpyxl
except ImportError:
    openpyxl = None

KO_SEND_SL = ("Initial: SL", "Echo: SL")
INS1_PATTERN = re.compile(r"\[INS1\]", re.I)
NAMETOKEN_PATTERN = re.compile(r"\(\(\s*nametoken\s*\)\)", re.I)

KO_COLUMNS = [
    "Stream Name",
    "Creative Name",
    "MLR Number",
    "WF Job Number (Billcode)",
    "Send",
    "Subject Lines + Preheaders",
    "Personalization",
    "Keycode 4",
]


def decode_csv_bytes(raw):
    """Decode CSV bytes using the best-fit encoding (Excel often saves as cp1252, not UTF-8)."""
    if not isinstance(raw, bytes):
        return raw if isinstance(raw, str) else str(raw)

    if raw.startswith(b"\xff\xfe"):
        return raw.decode("utf-16-le")
    if raw.startswith(b"\xfe\xff"):
        return raw.decode("utf-16-be")

    best_text = None
    best_replacements = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
        except UnicodeDecodeError:
            continue
        replacements = text.count("\ufffd")
        if best_replacements is None or replacements < best_replacements:
            best_replacements = replacements
            best_text = text

    return best_text if best_text is not None else raw.decode("utf-8", errors="replace")


def normalize_sl(text):
    """Normalize subject lines for strict comparison (nametoken/case/whitespace only)."""
    value = (text or "").strip()
    value = NAMETOKEN_PATTERN.sub("[INS1]", value)
    value = INS1_PATTERN.sub("[INS1]", value)
    value = value.replace("\u2019", "'").replace("\u2018", "'")
    value = re.sub(r"\s+", " ", value)
    return value.casefold()


def normalize_sl_loose(text):
    """Loose normalization for detecting encoding-only differences (dash/? variants)."""
    value = normalize_sl(text)
    dash_chars = "\u2010\u2011\u2012\u2013\u2014\u2015\u2212\ufe58\ufe63\uff0d"
    value = re.sub(f"[{re.escape(dash_chars)}]", "-", value)
    while re.search(r"(\w)\?(\w)", value):
        value = re.sub(r"(\w)\?(\w)", r"\1-\2", value)
    value = value.replace("\ufffd", "-")
    return value


def classify_subject_status(export_subject, ko_subject):
    """Return Match, Special character mismatch, or Subject mismatch."""
    export_text = (export_subject or "").strip()
    ko_text = (ko_subject or "").strip()
    if not export_text and not ko_text:
        return "Match"
    if bool(export_text) != bool(ko_text):
        return "Subject mismatch"
    if normalize_sl(export_subject) == normalize_sl(ko_subject):
        return "Match"
    if normalize_sl_loose(export_subject) == normalize_sl_loose(ko_subject):
        return "Special character mismatch"
    return "Subject mismatch"


def stream_id_to_keycode_stream(c_stream_id):
    match = re.match(r"S0?(\d+)", (c_stream_id or "").strip(), re.I)
    if match:
        return f"stream{int(match.group(1))}"
    return None


def build_keycode4(c_stream_id, c_creative_id):
    stream = stream_id_to_keycode_stream(c_stream_id)
    creative = (c_creative_id or "").strip().lower()
    if stream and creative:
        return f"{stream}|{creative}"
    return None


def parse_em_number(c_order_id):
    """Extract EM sequence number from c_order_id (e.g. EM03 -> 3)."""
    match = re.search(r"EM(\d+)", c_order_id or "", re.I)
    return int(match.group(1)) if match else None


def em_num_to_creative_idx(em_num):
    """Map EM number to creative block index: EM01/02 -> 0, EM03/04 -> 1, etc."""
    if em_num is None or em_num < 1:
        return None
    return (em_num - 1) // 2


def order_to_send(c_order_id):
    em_num = parse_em_number(c_order_id)
    if em_num is None:
        return None
    return "Initial: SL" if em_num % 2 == 1 else "Echo: SL"


def export_sl_display(subject, has_nametoken):
    """Format export subject the way KO expects in output."""
    text = (subject or "").strip()
    if has_nametoken:
        text = NAMETOKEN_PATTERN.sub("", text).strip()
        return f"[INS1] {text}".strip()
    return text


def _find_creative_details_sheet(workbook):
    for name in workbook.sheetnames:
        lower = name.lower().replace("_", " ")
        if "creative" in lower and "detail" in lower:
            return workbook[name]
    for name in workbook.sheetnames:
        if "creative" in name.lower():
            return workbook[name]
    return None


def _xlsx_to_ko_csv_text(file_bytes):
    """Read Creative Details tab from KO xlsx and return CSV text for parsing."""
    if openpyxl is None:
        raise ImportError("openpyxl required for .xlsx KO files. Run: py -3 -m pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = _find_creative_details_sheet(wb)
    if sheet is None:
        names = ", ".join(wb.sheetnames)
        wb.close()
        raise ValueError(f"Creative Details tab not found. Sheets: {names}")

    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(["" if v is None else str(v).strip() for v in row])
    wb.close()
    return buf.getvalue()


def parse_ko_document(file_obj):
    """Parse KO Creative Details from .xlsx or .csv."""
    if hasattr(file_obj, "read"):
        raw = file_obj.read()
    else:
        raw = file_obj

    if isinstance(raw, bytes):
        if raw[:2] == b"PK":
            raw = _xlsx_to_ko_csv_text(raw)
        else:
            raw = decode_csv_bytes(raw)
    elif not isinstance(raw, str):
        raw = str(raw)

    return _parse_ko_csv_text(raw)


def _flush_ko_block(block, records, ctx):
    """Emit SL records from a parsed creative block."""
    if not block:
        return
    jn = block.get("jn", "")
    keycode4 = block.get("keycode4", "")
    if not jn or not keycode4:
        return
    for entry in block.get("sl_entries", []):
        subject = entry["subject"]
        personalization = entry.get("personalization", "")
        records.append(
            {
                "stream_name": ctx["stream"],
                "creative_name": ctx["creative"],
                "mlr_number": ctx["mlr"],
                "jn": jn,
                "send": entry["send"],
                "subject": subject,
                "subject_normalized": normalize_sl(subject),
                "personalization": personalization,
                "keycode4": keycode4,
                "has_nametoken": "[INS1]" in subject.upper() or personalization.lower() == "yes",
            }
        )


def _parse_ko_csv_text(raw):
    """Parse KO creative-details CSV into SL rows."""
    reader = csv.reader(io.StringIO(raw))
    rows = list(reader)
    if not rows:
        return []

    header = rows[0]
    col_index = {name.strip(): idx for idx, name in enumerate(header)}

    def get(row, name, default=""):
        idx = col_index.get(name)
        if idx is None or idx >= len(row):
            return default
        return (row[idx] or "").strip()

    ctx = {"stream": "", "creative": "", "mlr": ""}
    current_block = None
    pending_jn = ""
    records = []

    for row in rows[1:]:
        stream = get(row, "Stream Name")
        creative = get(row, "Creative Name")
        mlr = get(row, "MLR Number")
        jn = get(row, "WF Job Number (Billcode)")
        send = get(row, "Send")
        subject = get(row, "Subject Lines + Preheaders")
        personalization = get(row, "Personalization")
        keycode4 = get(row, "Keycode 4").strip().lower()

        if stream:
            ctx["stream"] = stream
        if creative:
            ctx["creative"] = creative.replace("\n", " ").strip()
        if mlr:
            ctx["mlr"] = mlr

        # New creative block starts on Initial: SL with Keycode 4.
        if send == "Initial: SL" and keycode4:
            _flush_ko_block(current_block, records, ctx)
            current_block = {
                "jn": jn or pending_jn,
                "keycode4": keycode4,
                "sl_entries": [],
            }
            pending_jn = ""
        elif jn:
            if current_block is not None:
                current_block["jn"] = jn
            else:
                pending_jn = jn

        if send in KO_SEND_SL and current_block is not None:
            current_block["sl_entries"].append(
                {
                    "send": send,
                    "subject": subject,
                    "personalization": personalization,
                }
            )

    _flush_ko_block(current_block, records, ctx)
    return records


def keycode4_creative(keycode4):
    """Return the creative segment after stream| in Keycode 4."""
    parts = (keycode4 or "").split("|", 1)
    return parts[1] if len(parts) == 2 else (keycode4 or "")


def _ko_slot_has_sl(ko_row):
    return bool((ko_row.get("subject") or "").strip())


def _make_empty_ko_slot(stream, creative, send, template_row=None):
    """Placeholder for an expected EM slot with no SL content in the KO doc."""
    keycode4 = f"{stream}|{creative}"
    base = template_row or {}
    return {
        "stream_name": base.get("stream_name", ""),
        "creative_name": base.get("creative_name", ""),
        "mlr_number": base.get("mlr_number", ""),
        "jn": base.get("jn", ""),
        "send": send,
        "subject": "",
        "subject_normalized": "",
        "personalization": "",
        "keycode4": keycode4,
        "has_nametoken": False,
    }


def build_ko_stream_em_index(ko_rows):
    """
    Index ALL expected KO EM slots by (stream, em_num) in Creative Details block order.
    Within each stream: EM01=1st creative Initial, EM02=1st creative Echo, EM03=2nd Initial, etc.
    Missing or blank Echo/Initial rows still occupy their EM slot.
    """
    stream_creative_order = {}
    stream_creative_rows = {}

    for row in ko_rows:
        keycode4 = (row.get("keycode4") or "").strip().lower()
        parts = keycode4.split("|", 1)
        if len(parts) != 2:
            continue
        stream, creative = parts[0].strip(), parts[1].strip()
        if stream not in stream_creative_order:
            stream_creative_order[stream] = []
        if creative not in stream_creative_order[stream]:
            stream_creative_order[stream].append(creative)
        stream_creative_rows.setdefault((stream, creative), {})[row["send"]] = row

    by_stream_em = {}

    for stream, creatives in stream_creative_order.items():
        em_num = 1
        for creative in creatives:
            rows_for_creative = stream_creative_rows.get((stream, creative), {})
            template = rows_for_creative.get("Initial: SL") or rows_for_creative.get("Echo: SL")
            for send in KO_SEND_SL:
                ko_row = rows_for_creative.get(send)
                if ko_row is None:
                    ko_row = _make_empty_ko_slot(stream, creative, send, template)
                by_stream_em[(stream, em_num)] = ko_row
                em_num += 1

    return by_stream_em


def _find_ko_slot(by_stream_em, export_stream, em_num, creative):
    """
    Find KO slot for export row: primary (stream, em_num), then stream-remap by EM + creative.
    Returns (ko_row, stream_aligned).
    """
    creative = (creative or "").strip().lower()
    if export_stream and em_num:
        slot = by_stream_em.get((export_stream, em_num))
        if slot and keycode4_creative(slot["keycode4"]).lower() == creative:
            return slot, True
        if slot:
            slot = None

    if em_num and creative:
        for (ko_stream, ko_em), candidate in by_stream_em.items():
            if ko_em != em_num:
                continue
            if keycode4_creative(candidate["keycode4"]).lower() != creative:
                continue
            return candidate, export_stream == ko_stream
    return None, False


def _validation_record(export_row, ko_row, status, em_num=None):
    record = {
        "jn": export_row["jn"] if export_row else (ko_row or {}).get("jn", ""),
        "send": (export_row or ko_row)["send"],
        "keycode4": (export_row or ko_row).get("keycode4", ""),
        "export_subject": (export_row or {}).get("subject", ""),
        "ko_subject": (ko_row or {}).get("subject", ""),
        "status": status,
        "c_stream_id": (export_row or {}).get("c_stream_id", ""),
        "c_order_id": (export_row or {}).get("c_order_id", ""),
        "c_creative_id": (export_row or {}).get("c_creative_id", ""),
        "em_num": em_num if em_num is not None else (export_row or {}).get("em_num"),
    }
    if export_row and ko_row:
        if ko_row.get("jn") and ko_row["jn"] != export_row["jn"]:
            record["ko_jn"] = ko_row["jn"]
        if ko_row.get("keycode4") and ko_row["keycode4"] != export_row.get("keycode4"):
            record["ko_keycode4"] = ko_row["keycode4"]
        ko_stream = (ko_row.get("keycode4") or "").split("|")[0]
        export_stream = stream_id_to_keycode_stream(export_row.get("c_stream_id"))
        if ko_stream and export_stream and ko_stream != export_stream:
            record["ko_stream"] = ko_stream
    elif ko_row:
        record["keycode4"] = ko_row.get("keycode4", record["keycode4"])
        record["send"] = ko_row.get("send", record["send"])
        if not record["c_creative_id"]:
            record["c_creative_id"] = keycode4_creative(ko_row.get("keycode4", ""))
    return record


def validate_export_against_ko(export_rows, ko_rows):
    """
    Compare export SL rows to KO document using stream + EM order + creative + subject.
    Every EM slot is indexed in the KO doc (including blank/missing Echo SL).
    """
    by_stream_em = build_ko_stream_em_index(ko_rows)

    matched = []
    mismatches = []
    export_only = []
    ko_only = []
    matched_slot_keys = set()

    for export_row in export_rows:
        export_stream = stream_id_to_keycode_stream(export_row.get("c_stream_id"))
        em_num = export_row.get("em_num")
        if em_num is None:
            em_num = parse_em_number(export_row.get("c_order_id"))

        creative = (
            export_row.get("c_creative_id") or keycode4_creative(export_row.get("keycode4"))
        ).strip().lower()
        export_has = _ko_slot_has_sl(export_row)

        ko_row, stream_aligned = _find_ko_slot(by_stream_em, export_stream, em_num, creative)

        if ko_row is None:
            export_only.append(
                _validation_record(
                    export_row,
                    None,
                    "Not in KO doc",
                    em_num=em_num,
                )
            )
            continue

        ko_has = _ko_slot_has_sl(ko_row)
        ko_stream = (ko_row.get("keycode4") or "").split("|")[0]
        slot_key = (em_num, creative, export_row["send"])

        if export_has and not ko_has:
            mismatches.append(
                _validation_record(export_row, ko_row, "Missing in KO", em_num=em_num)
            )
            continue

        if not export_has and ko_has:
            mismatches.append(
                _validation_record(export_row, ko_row, "Missing in export", em_num=em_num)
            )
            continue

        if not export_has and not ko_has:
            matched_slot_keys.add(slot_key)
            continue

        status = classify_subject_status(export_row["subject"], ko_row["subject"])
        record = _validation_record(export_row, ko_row, status, em_num=em_num)

        if status == "Match":
            matched.append(record)
            matched_slot_keys.add(slot_key)
        else:
            mismatches.append(record)

    for (stream, em_num), ko_row in by_stream_em.items():
        if not _ko_slot_has_sl(ko_row):
            continue
        creative = keycode4_creative(ko_row["keycode4"]).lower()
        slot_key = (em_num, creative, ko_row["send"])
        if slot_key in matched_slot_keys:
            continue
        ko_only.append(
            {
                "jn": ko_row.get("jn", ""),
                "send": ko_row["send"],
                "keycode4": ko_row.get("keycode4", ""),
                "export_subject": "",
                "ko_subject": ko_row["subject"],
                "status": "Missing in export",
                "c_stream_id": "",
                "c_order_id": f"EM{em_num:02d}",
                "c_creative_id": creative,
                "em_num": em_num,
                "ko_jn": ko_row.get("jn", ""),
            }
        )

    missing_in_export = [
        r for r in mismatches + ko_only if r["status"] == "Missing in export"
    ]
    missing_in_ko = [r for r in mismatches if r["status"] == "Missing in KO"]

    return {
        "matched": matched,
        "mismatches": mismatches,
        "export_only": export_only,
        "ko_only": ko_only,
        "match_count": len(matched),
        "mismatch_count": len(mismatches),
        "export_only_count": len(export_only),
        "ko_only_count": len(ko_only),
        "missing_in_export_count": len(missing_in_export),
        "missing_in_ko_count": len(missing_in_ko),
    }


def build_ko_aligned_csv(export_rows):
    """Build CSV aligned to KO document SL columns."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(KO_COLUMNS)
    last_jn = None
    for row in export_rows:
        include_header_fields = row["jn"] != last_jn
        writer.writerow(
            [
                "" if not include_header_fields else f"Stream from export ({row['c_stream_id']})",
                "" if not include_header_fields else row.get("action_name", ""),
                "NA",
                row["jn"] if include_header_fields else "",
                row["send"],
                row["subject_ko_format"],
                row["personalization"],
                row["keycode4"],
            ]
        )
        last_jn = row["jn"]
    return buf.getvalue()


def build_validation_csv(validation, export_label="Subject (Campaign Export)", ko_label="Subject (KO Creative Details doc)"):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "Status",
            "Stream",
            "EM Order",
            "c_creative_id",
            "JN",
            "Send",
            "Keycode 4",
            export_label,
            ko_label,
            "KO JN",
            "KO Keycode 4",
        ]
    )
    for group in ("matched", "mismatches", "export_only", "ko_only"):
        for row in validation[group]:
            writer.writerow(
                [
                    row["status"],
                    row.get("c_stream_id", ""),
                    row.get("c_order_id", ""),
                    row.get("c_creative_id", ""),
                    row["jn"],
                    row["send"],
                    row["keycode4"],
                    row["export_subject"],
                    row["ko_subject"],
                    row.get("ko_jn", ""),
                    row.get("ko_keycode4", ""),
                ]
            )
    return buf.getvalue()
